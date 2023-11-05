import os
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

import numpy as np
import pandas as pd
import openpyxl

def get_data(train_x):
    record = []
    count = 0
    if MODE==1:
        for i in range(len(train_x)):
            if (train_x[i][feature_name[0]]==0)and(count<TrainX_size):
                count = count+1
                for k in range(data_size):
                    record.append(train_x[i])
    if MODE==2:
        for i in range(len(train_x)):
            if (train_x[i][feature_name[0]]!=0)and(count<TrainX_size):
                count = count+1
                for k in range(data_size):
                    record.append(train_x[i])
    if MODE==3:
        for i in range(len(train_x)):
            if (train_x[i][feature_name[0]]==0)and(train_x[i][feature_name[1]]==0)and(count<TrainX_size):
                count = count+1
                for k in range(data_size):
                    record.append(train_x[i])
    if MODE==4:
        for i in range(len(train_x)):
            if (train_x[i][feature_name[0]]!=0)and(train_x[i][feature_name[1]]!=0)and(count<TrainX_size):
                count = count+1
                for k in range(data_size):
                    record.append(train_x[i])
    if MODE==5:
        for i in range(len(train_x)):
            if (train_x[i][feature_name[0]]==0)and(train_x[i][feature_name[1]]!=0)and(count<TrainX_size):
                count = count+1
                for k in range(data_size):
                    record.append(train_x[i])
    if MODE==6:
        for i in range(len(train_x)):
            if (train_x[i][feature_name[0]]!=0)and(train_x[i][feature_name[1]]==0)and(count<TrainX_size):
                count = count+1
                for k in range(data_size):
                    record.append(train_x[i])
    print('get_data len: ',len(record))
    return record

def handel_input():
    with np.load(TARGET_DATA) as f:
            train_x, _, _, _ = [f['arr_%d' % i] for i in range(len(f.files))]  
    record = get_data(train_x)
    wINPUT = openpyxl.load_workbook(ORIGINAL_INPUT_PATH, data_only=True)
    s1 = wINPUT['工作表1']
    for j in range(TrainX_size):
        for k in range(data_size):
            record[(j*data_size)+k] = np.insert(record[(j*data_size)+k],0, s1.cell(k+1,2).value)
    return record

def handel_label():
    label = []
    wINPUT = openpyxl.load_workbook(ORIGINAL_INPUT_PATH, data_only=True)
    s1 = wINPUT['工作表1']
    for i in range(TrainX_size):
        for j in range(len(class_name)):
            for k in range(round):
                tmp = [j]
                label.append(tmp)
    count = 0
    for j in range(TrainX_size):
        for k in range(data_size):
            label[count].append(s1.cell(k+1,1).value)
            count = count+1
    return label

data_name = ['feature16_zero','feature37_zero','feature16_not_zero','feature37_not_zero']
mode = [1,1,2,2]
feature = [16,37,16,37]

# data_name = ['feature16_not_zero_37_not_zero','feature16_not_zero_37_zero','feature16_zero_37_not_zero','feature16_zero_37_zero']
# mode = [4,6,5,3]
# feature = [16,37]

class_name = [0.01,0.05,0.1,0.5,1,5,10,100,1000]
round = 35
data_size = round*len(class_name)
TrainX_size = 4000

for i in range(4):
    feature_name = [feature[i]]
    # feature_name = feature
    DATA = '9_35_100/'+data_name[i]
    ORIGINAL_INPUT_PATH = 'original_data/'+DATA+'.xlsx'
    INPUT_PATH = 'data/leak2/'+data_name[i]+'.csv'
    LABEL_PATH = 'data/leak2/'+data_name[i]+'_label.csv'
    TARGET_DATA = 'original_data/target_data/feature100_target_data.npz'
    MODE = mode[i]

    print('run: ',i)
    print('data name: ',DATA)
    print('feature name: ',feature_name[0])
    # print('feature name: ',feature_name[0],feature_name[1])
    print('mode: ',MODE)
    Record = handel_input()
    Label = handel_label()

    # print('len(Record): ',len(Record))
    # print('len(Label): ',len(Label))

    ##################### save input #####################

    input_df = pd.DataFrame(Record)
    input_df.to_csv(INPUT_PATH,index=False,header=None)

    ##################### save label #####################

    label_df = pd.DataFrame(Label)
    label_df.to_csv(LABEL_PATH,index=False,header=None)

    tmp1 = np.array(Record)
    tmp2 = np.array(Label)
    print('Record: ',tmp1.shape)
    print('Label: ',tmp2.shape)

# data_name = ['feature16_zero','feature37_zero','feature16_not_zero','feature37_not_zero']
# mode = [1,1,2,2]
# feature = [16,37,16,37]

data_name = ['feature16_not_zero_37_not_zero','feature16_not_zero_37_zero','feature16_zero_37_not_zero','feature16_zero_37_zero']
mode = [4,6,5,3]
feature = [16,37]

class_name = [0.01,0.05,0.1,0.5,1,5,10,100,1000]
round = 35
data_size = round*len(class_name)
TrainX_size = 4000

for i in range(4):
    # feature_name = [feature[i]]
    feature_name = feature
    DATA = '9_35_100/'+data_name[i]
    ORIGINAL_INPUT_PATH = 'original_data/'+DATA+'.xlsx'
    INPUT_PATH = 'data/leak2/'+data_name[i]+'.csv'
    LABEL_PATH = 'data/leak2/'+data_name[i]+'_label.csv'
    TARGET_DATA = 'original_data/target_data/feature100_target_data.npz'
    MODE = mode[i]

    print('run: ',i)
    print('data name: ',DATA)
    # print('feature name: ',feature_name[0])
    print('feature name: ',feature_name[0],feature_name[1])
    print('mode: ',MODE)
    Record = handel_input()
    Label = handel_label()

    ##################### save input #####################

    input_df = pd.DataFrame(Record)
    input_df.to_csv(INPUT_PATH,index=False,header=None)

    ##################### save label #####################

    label_df = pd.DataFrame(Label)
    label_df.to_csv(LABEL_PATH,index=False,header=None)

    tmp1 = np.array(Record)
    tmp2 = np.array(Label)
    print('Record: ',tmp1.shape)
    print('Label: ',tmp2.shape)


def handel_input():
    with np.load(TARGET_DATA) as f:
            train_x, _, _, _ = [f['arr_%d' % i] for i in range(len(f.files))]  
    record = get_data(train_x)
    wINPUT_leak = openpyxl.load_workbook(ORIGINAL_INPUT_PATH, data_only=True)
    s1 = wINPUT_leak['工作表1']
    for j in range(TrainX_size):
        for k in range(data_size):
            record[(j*data_size)+k] = np.insert(record[(j*data_size)+k],0, s1.cell(k+1,2).value)
    return record

def handel_label():
    label = []
    wINPUT_acc = openpyxl.load_workbook(ORIGINAL_INPUT_PATH, data_only=True)
    s1 = wINPUT_acc['工作表1']
    for i in range(TrainX_size):
        for j in range(len(class_name)):
            for k in range(round):
                tmp = [j]
                label.append(tmp)
    count = 0
    for j in range(TrainX_size):
        for k in range(data_size):
            label[count].append(s1.cell(k+1,1).value)
            count = count+1
    return label