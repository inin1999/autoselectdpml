import os
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

import numpy as np
import pandas as pd
import openpyxl
from sklearn.model_selection import train_test_split

class_name = [0.01,0.05,0.1,0.5,1,5,10,100,1000]
round = 14
data_size = round*len(class_name)
len_train_x = 10000

DATASET = 'cifar100'
DATA = DATASET+'/9_14_/all'
ORIGINAL_INPUT_PATH = 'original_data/'+DATA+'.xlsx'
INPUT_PATH = 'data/'+DATA+'.csv'
LABEL_PATH = 'data/'+DATA+'_label.csv'
TARGET_DATA = 'original_data/'+DATASET+'/target_data/target_data.npz'

def get_data(train_x):
    record = []
    for i in range(len(train_x)):
        for k in range(data_size):
            record.append(train_x[i])
    return record

def handel_input():
    with np.load(TARGET_DATA) as f:
            train_x, _, _, _ = [f['arr_%d' % i] for i in range(len(f.files))]    
    record = get_data(train_x)
    print('len(record):', len(record))
    print('len(train_x[0]:', len(train_x[0]))
    wINPUT = openpyxl.load_workbook(ORIGINAL_INPUT_PATH, data_only=True)
    s1 = wINPUT['工作表1']
    for j in range(len(train_x)):
        for k in range(data_size):
            record[(j*data_size)+k] = np.insert(record[(j*data_size)+k],0, s1.cell(k+1,1).value)
            record[(j*data_size)+k] = np.insert(record[(j*data_size)+k],1, s1.cell(k+1,2).value)
    print('len(record[0]): ',len(record[0]))
    print('len(record[len(record)-1]): ',len(record[len(record)-1]))
    return record

def handel_label():
    label = []
    count = 0
    for i in range(len_train_x):
        for j in range(len(class_name)):
            for k in range(round):
                count = count+1
                print('count label: ',count)
                label.append(j)
    return label

print('data name: ',DATA)
Record = handel_input()
Label = handel_label()

print('len(Record):',len(Record))
print('len(Label):',len(Label))

##################### save input #####################

input_df = pd.DataFrame(Record)
input_df.to_csv(INPUT_PATH,index=False,header=None)

##################### save label #####################

label_df = pd.DataFrame(Label)
label_df.to_csv(LABEL_PATH,index=False,header=None)