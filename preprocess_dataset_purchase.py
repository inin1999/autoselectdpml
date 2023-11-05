import os
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

import numpy as np
import pandas as pd
import openpyxl
from sklearn.model_selection import train_test_split

class_name = [0.01,0.05,0.1,0.5,1,5,10,100,1000]
feature_name = [16,37]
round = 35
data_size = round*len(class_name)
TrainX_size = 4000

DATASET = 'purchase'
DATA = DATASET+'/9_35_100/feature16_zero'
ORIGINAL_INPUT_PATH = 'original_data/'+DATA+'.xlsx'
INPUT_PATH = 'data/'+DATA+'2.csv'
LABEL_PATH = 'data/'+DATA+'2_label.csv'
TARGET_DATA = 'original_data/'+DATASET+'/target_data/target_data.npz'
MODE = 1

def get_data(train_x):
    record = []
    count = 0
    if MODE==1:
        for i in range(len(train_x)):
            if (train_x[i][feature_name[0]]==0)and(count<TrainX_size):
                count = count+1
                for k in range(data_size):
                    record.append(train_x[i])
    if MODE==2:
        for i in range(len(train_x)):
            if (train_x[i][feature_name[0]]!=0)and(count<TrainX_size):
                count = count+1
                for k in range(data_size):
                    record.append(train_x[i])
    if MODE==3:
        for i in range(len(train_x)):
            if (train_x[i][feature_name[0]]==0)and(train_x[i][feature_name[1]]==0)and(count<TrainX_size):
                count = count+1
                for k in range(data_size):
                    record.append(train_x[i])
    if MODE==4:
        for i in range(len(train_x)):
            if (train_x[i][feature_name[0]]!=0)and(train_x[i][feature_name[1]]!=0)and(count<TrainX_size):
                count = count+1
                for k in range(data_size):
                    record.append(train_x[i])
    if MODE==5:
        for i in range(len(train_x)):
            if (train_x[i][feature_name[0]]==0)and(train_x[i][feature_name[1]]!=0)and(count<TrainX_size):
                count = count+1
                for k in range(data_size):
                    record.append(train_x[i])
    if MODE==6:
        for i in range(len(train_x)):
            if (train_x[i][feature_name[0]]!=0)and(train_x[i][feature_name[1]]==0)and(count<TrainX_size):
                count = count+1
                for k in range(data_size):
                    record.append(train_x[i])
    return record

def handel_input():
    with np.load(TARGET_DATA) as f:
            train_x, _, _, _ = [f['arr_%d' % i] for i in range(len(f.files))]    
    record = get_data(train_x)
    print('len(record):', len(record))
    print('len(train_x[0]:', len(train_x[0]))
    wINPUT = openpyxl.load_workbook(ORIGINAL_INPUT_PATH, data_only=True)
    s1 = wINPUT['工作表1']
    for j in range(TrainX_size):
        for k in range(data_size):
            record[(j*data_size)+k] = np.insert(record[(j*data_size)+k],0, s1.cell(k+1,1).value)
            record[(j*data_size)+k] = np.insert(record[(j*data_size)+k],1, s1.cell(k+1,2).value)
    print('len(record[0]): ',len(record[0]))
    print('len(record[len(record)-1]): ',len(record[len(record)-1]))
    return record

def handel_label():
    label = []
    count = 0
    for i in range(TrainX_size):
        for j in range(len(class_name)):
            for k in range(round):
                count = count+1
                print('count label: ',count)
                label.append(j)
    return label

print('data name: ',DATA)
print('feature name: ',feature_name[0],' ',feature_name[1])
print('mode: ',MODE)
Record = handel_input()
Label = handel_label()

print(len(Record))
print(len(Label))

##################### save input #####################

input_df = pd.DataFrame(Record)
input_df.to_csv(INPUT_PATH,index=False,header=None)

##################### save label #####################

label_df = pd.DataFrame(Label)
label_df.to_csv(LABEL_PATH,index=False,header=None)